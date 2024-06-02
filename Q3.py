# -*- coding: gbk -*-
import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader
from torchvision import transforms
import os
from PIL import Image
import numpy as np
import matplotlib.pyplot as plt
import cv2
from openpyxl import load_workbook
# image_ori = cv2.imread(os.path.join('4_Recognize','test_set','w01790.jpg'))
# cv2.imshow('Image', image_ori)

# cv2.waitKey(0)
# # �رմ���
# cv2.destroyAllWindows()

picture_num = 6148 # ����ģ��ʱ���ܹ������ͼƬ����
num_epochs = 25 # ����ģ��ʱ����������
outputs_threshold = 0.36 # ����ģ��ʱ�������ֵ�����ֵ��ʼ����0.05�ݼ�

class UNet(nn.Module): # ���ĵĻ���ѧϰ����
    def __init__(self):
        super(UNet, self).__init__()
        # ������
        self.enc_conv1 = self.encoder_block(1, 8)
        self.enc_conv2 = self.encoder_block(8, 16)
        self.enc_conv3 = self.encoder_block(16, 32)
        self.enc_conv4 = self.encoder_block(32, 128)

        self.pool = nn.MaxPool2d(kernel_size=2, stride=2)
        # ������
        self.upconv4 = nn.ConvTranspose2d(128, 64, kernel_size=2, stride=2)
        self.dec_conv4 = self.decoder_block(192, 64)

        self.upconv3 = nn.ConvTranspose2d(64, 32, kernel_size=2, stride=2)
        self.dec_conv3 = self.decoder_block(64, 32)

        self.upconv2 = nn.ConvTranspose2d(32, 16, kernel_size=2, stride=2)
        self.dec_conv2 = self.decoder_block(32, 16)
        
        self.upconv1 = nn.ConvTranspose2d(16, 8, kernel_size=2, stride=2)
        self.dec_conv1 = self.decoder_block(16, 8)
        # ���ղ�
        self.final_conv = nn.Conv2d(8, 1, kernel_size=1)

    def encoder_block(self, in_channels, out_channels):
        block = nn.Sequential(
            nn.Conv2d(in_channels, out_channels, kernel_size=3, padding=1),
            nn.ReLU(inplace=True),
            nn.Conv2d(out_channels, out_channels, kernel_size=3, padding=1),
            nn.ReLU(inplace=True),
        )
        return block
    def decoder_block(self, in_channels, out_channels):
        block = nn.Sequential(
            nn.Conv2d(in_channels, out_channels, kernel_size=3, padding=1),
            nn.ReLU(inplace=True),
            nn.Conv2d(out_channels, out_channels, kernel_size=3, padding=1),
            nn.ReLU(inplace=True),
        )
        return block
    def forward(self, x):
        # ����
        enc1 = self.enc_conv1(x)
        pool1 = self.pool(enc1)
        
        enc2 = self.enc_conv2(pool1)
        pool2 = self.pool(enc2)
        
        enc3 = self.enc_conv3(pool2)
        pool3 = self.pool(enc3)
        
        enc4 = self.enc_conv4(pool3)
        pool4 = self.pool(enc4)
        
        # ����
        up4 = self.upconv4(pool4)
        merge4 = torch.cat([up4, enc4], dim=1)
        dec4 = self.dec_conv4(merge4)
        
        up3 = self.upconv3(dec4)
        merge3 = torch.cat([up3, enc3], dim=1)
        dec3 = self.dec_conv3(merge3)
        
        up2 = self.upconv2(dec3)
        merge2 = torch.cat([up2, enc2], dim=1)
        dec2 = self.dec_conv2(merge2)
        
        up1 = self.upconv1(dec2)
        merge1 = torch.cat([up1, enc1], dim=1)
        dec1 = self.dec_conv1(merge1)
        # ���ղ�
        return torch.sigmoid(self.final_conv(dec1))

class UnlabeledImageDataset(Dataset): # û�б�ǩ�����ݼ�
    def __init__(self, image_dir):
        self.image_dir = image_dir
        self.transform = transforms.ToTensor()
        self.images = os.listdir(image_dir)

    def __len__(self):
        return len(self.images)

    def __getitem__(self, idx):
        img_name = self.images[idx]
        img_path = os.path.join(self.image_dir, img_name)

        image = Image.open(img_path).convert('L') # �ԻҶ�ͼ���ȡ
        
        image = image.resize((128, 128)) # ѹ���ߴ�
        
        image = self.transform(image) # ת��Ϊ����
        
        return image, img_name
image_dir = os.path.join('4_Recognize','test_set')  # ����ͼƬĿ¼·��
    
def find_value_row(filename, sheetname, value):
    # ��Excel�ļ�
    wb = load_workbook(filename)
    
    # ѡ������
    ws = wb[sheetname]
    i = 1
    # ������һ�У�����ֵ���ڵ��к�
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        if row[0] == value:
            print(i)
            return i  # ֱ�ӷ����к�
        else:
            i += 1
    # ���ֵδ�ҵ�������None
    return None
unlabeled_dataset = UnlabeledImageDataset(image_dir) # ��������

# �������ݼ�����
data_loader = DataLoader(unlabeled_dataset, batch_size=1, shuffle=False)

model_path = 'unet_model_6148_25_att.pth' # ����ģ��
model_state_dict = torch.load(model_path)
model = UNet()
model.load_state_dict(model_state_dict)
    
model.eval() # ����Ϊ����ģʽ

with torch.no_grad():
    for images, img_name in data_loader:
                        
        image_name = img_name[0]
            
        img_path = os.path.join(image_dir, image_name) # ��ȡͼƬ·��
        print(img_path)
        image_ori = cv2.imread(img_path) # ��ȡͼƬ��ԭʼͼƬ��
            
        height_original, width_original, _ = image_ori.shape # ע�⣺cv2��ͼ��ߴ��˳����(�߶�, ���)

        width_scale = width_original / 128 # ��ȡԭʼͼƬ�����ű���
            
        height_scale = height_original / 128 # ��ȡԭʼͼƬ�����ű���
            
        outputs = model(images)  # ʹ��ģ�ͽ���Ԥ��
            
        preds = (outputs >= outputs_threshold).float()  # Ӧ����ֵ
            
        image_data = preds.squeeze().cpu().numpy()  # ת��ΪNumPy���飬���image_data��֮�����չʾ����
            
        preds_np = preds.squeeze().cpu().numpy() # ת��ΪNumPy����
            
        preds_np = (preds_np * 255).astype(np.uint8)  # ת��Ϊ8-bit���飬ֵΪ0��255
            
        contours, _ = cv2.findContours(preds_np, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE) # ��ȡ��С��Ӿ���
            
        character_rect_list = []
            
        for contour in contours: # �������������ƾ���
                
            x, y, w, h = cv2.boundingRect(contour) # ��С��Ӿ���
            if (w * h) > 22:  # ֻ�е��������ָ����Сʱ�Ż���
                x1_original = int(x * width_scale) - 1 # ���Ͻ����꣬��ӳ�䵽ԭʼ��ͼ��������
                y1_original = int(y * height_scale) - 1 
                x2_original = int((x+w) * width_scale) + 1 # ���½�����
                y2_original = int((y+h) * height_scale) + 1
                    
                character_rect_list.append([x1_original,y1_original,x2_original,y2_original, 1.0])
                    
                
                if 1 == 2: # �Ƿ񱣴���к��ͼƬ

                    cropped_image = image_ori[y1_original:y2_original, x1_original:x2_original]
                    cropped_image_name = f"{image_name[:-4]}_{x1_original}_{y1_original}_{x2_original}_{y2_original}.jpg"

                    # ������ȡ��ͼ��
                    cropped_image_save_path = os.path.join('saved_crops', cropped_image_name)

                    cv2.imwrite(cropped_image_save_path, cropped_image)

                cv2.rectangle(image_ori, (x1_original, y1_original), (x2_original, y2_original), (0,0,255), thickness=1)  # ��ɫ��

        print(str(character_rect_list))

        if 1 == 2: # �ж��Ƿ񽫲��н��д��excel���

            wb = load_workbook(os.path.join("3_Test", "Test_results.xlsx"))
            ws = wb['Sheet1']
            # for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
            #     if row[0] == image_name:
            #         print(row[0])
            #         ws.cell(row[0].row, column=2, value=str(character_rect_list))
            row_number = find_value_row(os.path.join("3_Test", "Test_results.xlsx"), 'Sheet1', image_name)
            ws.cell(row = row_number, column = 2, value = str(character_rect_list)[1: -1]) 
            wb.save(os.path.join("3_Test", "Test_results.xlsx"))
        if 1 == 2: # ѡ���Ƿ񱣴�ͼƬ�����ƺ�ɫ���Σ�
                
            save_path = os.path.join('saved_crops','ori_pics',img_name[0]) # ƴ��ͼƬ����·��
            
            cv2.imwrite(save_path, image_ori) # ��������������ԭͼƬ�ϻ��ƾ��εĽ����
            
        if 1 == 1: # ѡ���Ƿ���ʾͼ��

            image_ori = cv2.cvtColor(image_ori, cv2.COLOR_BGR2RGB)

            plt.imshow(preds_np, cmap='gray') # ��ʾ2ֵ����Ԥ�ⷶΧ
            plt.axis('off')
            plt.show()
            
            plt.imshow(image_ori) # ��ԭͼƬ�ϻ��ƺ�ɫ���ε�ͼƬ
            plt.axis('off')
            plt.show()

            plt.imshow(image_data, cmap='gray') # ԭʼͼƬ
            plt.axis('off')
            plt.show()
                
            